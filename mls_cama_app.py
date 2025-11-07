import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
import os

# Install required package if not available
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--break-system-packages", "openpyxl"])
    import openpyxl
    from openpyxl import load_workbook

# Try to import windowId extractor
try:
    from cama_windowid_extractor import get_window_id
    EXTRACTOR_AVAILABLE = True
except ImportError:
    EXTRACTOR_AVAILABLE = False

# Set page configuration
st.set_page_config(
    page_title="MLS vs CAMA Comparison Tool",
    page_icon="🏠",
    layout="wide"
)

# Title and description
st.title("🏠 MLS vs CAMA Data Comparison Tool")
st.markdown("""
This application compares property data between MLS (Multiple Listing Service) and CAMA (Assessment) systems.
Upload your Excel files and configure the comparison settings below.
""")

# Sidebar for configuration
st.sidebar.header("⚙️ Configuration")

# File uploads
st.sidebar.subheader("📁 Upload Files")
mls_file = st.sidebar.file_uploader("Upload MLS Data (Excel)", type=['xlsx', 'xls'])
cama_file = st.sidebar.file_uploader("Upload CAMA Data (Excel)", type=['xlsx', 'xls'])

# Column mappings
st.sidebar.subheader("🔗 Column Mappings")
unique_id_mls = st.sidebar.text_input("MLS ID Column", "Parcel Number")
unique_id_cama = st.sidebar.text_input("CAMA ID Column", "PARID")

# Comparison settings
st.sidebar.subheader("⚖️ Comparison Settings")
numeric_tolerance = st.sidebar.number_input("Numeric Tolerance", value=0.01, format="%.4f")
skip_zero_values = st.sidebar.checkbox("Skip Zero Values", value=True)

# URL templates and WindowId
st.sidebar.subheader("🔗 Hyperlink Settings")

# Auto-extraction feature
if EXTRACTOR_AVAILABLE:
    use_auto_extract = st.sidebar.checkbox("🤖 Auto-extract WindowId", value=True, 
                                           help="Automatically get fresh windowId from CAMA system")
    
    if use_auto_extract:
        with st.sidebar.expander("🔐 CAMA Credentials (Optional)"):
            st.info("Credentials improve extraction success rate. Leave blank to try without login.")
            cama_username = st.text_input("Username", type="default", key="cama_user")
            cama_password = st.text_input("Password", type="password", key="cama_pass")
            
            if st.button("🔍 Extract WindowId Now"):
                with st.spinner("Extracting windowId from CAMA..."):
                    extracted_id = get_window_id(
                        username=cama_username if cama_username else None,
                        password=cama_password if cama_password else None,
                        fallback_id="638981240146803746"
                    )
                    if extracted_id:
                        st.session_state.extracted_window_id = extracted_id
                        st.success(f"✅ Extracted: {extracted_id}")
                    else:
                        st.error("❌ Extraction failed")
        
        # Use extracted or fallback
        if 'extracted_window_id' in st.session_state:
            window_id = st.session_state.extracted_window_id
            st.sidebar.success(f"✅ Using: {window_id}")
        else:
            window_id = "638981240146803746"
            st.sidebar.info(f"ℹ️  Using fallback: {window_id}")
    else:
        # Manual entry
        st.sidebar.info("💡 Get windowId from CAMA website URL")
        window_id = st.sidebar.text_input(
            "🔑 WindowId",
            "638981240146803746",
            help="Get this from the CAMA website URL"
        )
else:
    # No extractor available - manual only
    st.sidebar.info("""
    💡 **How to update WindowId:**
    1. Go to the CAMA website
    2. Search for any property
    3. Copy the windowId from the URL
    4. Paste it below
    """)
    window_id = st.sidebar.text_input(
        "🔑 WindowId (from CAMA system)",
        "638981240146803746",
        help="Get this from the CAMA website URL"
    )

# Build the URL template with the user's windowId
parcel_url_template = f"https://iasworld.starkcountyohio.gov/iasworld/Maintain/Transact.aspx?txtMaskedPin={{parcel_id}}&selYear=&userYear=&selJur=&chkShowHistory=False&chkShowChanges=&chkShowDeactivated=&PinValue={{parcel_id}}&pin=&trans_key=&windowId={window_id}&submitFlag=true&TransPopUp=&ACflag=False&ACflag2=False"

# Column comparisons to perform
COLUMNS_TO_COMPARE = [
    {'mls_col': 'Above Grade Finished Area', 'cama_col': 'SFLA'},
    {'mls_col': 'Bedrooms Total', 'cama_col': 'RMBED'},
    {'mls_col': 'Bathrooms Full', 'cama_col': 'FIXBATH'},
    {'mls_col': 'Bathrooms Half', 'cama_col': 'FIXHALF'},
]

COLUMNS_TO_COMPARE_SUM = [
    {'mls_col': 'Below Grade Finished Area', 'cama_cols': ['RECROMAREA', 'FINBSMTAREA', 'UFEATAREA']}
]

COLUMNS_TO_COMPARE_CATEGORICAL = [
    {
        'mls_col': 'Cooling',
        'cama_col': 'HEAT',
        'mls_check_contains': 'Central Air',
        'cama_expected_if_true': 1,
        'cama_expected_if_false': 0,
        'case_sensitive': False
    }
]

ADDRESS_COLUMNS = {
    'address': 'Address',
    'city': 'City',
    'state': 'State or Province',
    'zip': 'Postal Code'
}

# Utility functions
def format_zillow_url(address, city, state, zip_code):
    """Create a Zillow search URL from address components."""
    if pd.isna(address) or pd.isna(city) or pd.isna(zip_code):
        return None
    
    address_clean = str(address).strip()
    city_clean = str(city).strip()
    zip_clean = str(zip_code).strip().split('-')[0]
    
    address_clean = re.sub(r'\s+(Apt|Unit|#|Suite)\s*[\w-]*$', '', address_clean, flags=re.IGNORECASE)
    address_formatted = re.sub(r'[^\w\s-]', '', address_clean)
    address_formatted = re.sub(r'\s+', '-', address_formatted)
    
    city_formatted = re.sub(r'[^\w\s-]', '', city_clean)
    city_formatted = re.sub(r'\s+', '-', city_formatted)
    
    url_slug = f"{address_formatted}-{city_formatted}-OH-{zip_clean}_rb"
    return f"https://www.zillow.com/homes/{url_slug}/"

def values_equal(val1, val2, tolerance):
    """Check if two values are equal within tolerance."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')
        
        if pd.isna(num1) and pd.isna(num2):
            return True
        elif pd.isna(num1) != pd.isna(num2):
            return False
        else:
            return np.isclose(num1, num2, equal_nan=False, rtol=1e-9, atol=tolerance)
    except (ValueError, TypeError):
        str1 = str(val1).strip().lower() if pd.notna(val1) else ''
        str2 = str(val2).strip().lower() if pd.notna(val2) else ''
        return str1 == str2

def categorical_match(mls_val, cama_val, mapping, tolerance):
    """Check if a categorical MLS field matches expected CAMA value."""
    check_text = mapping.get('mls_check_contains', '')
    expected_if_true = mapping.get('cama_expected_if_true')
    expected_if_false = mapping.get('cama_expected_if_false')
    case_sensitive = mapping.get('case_sensitive', False)
    
    mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
    
    if not case_sensitive:
        mls_str = mls_str.lower()
        check_text = check_text.lower()
    
    text_found = check_text in mls_str
    expected_cama = expected_if_true if text_found else expected_if_false
    
    try:
        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
        expected_numeric = pd.to_numeric(expected_cama, errors='coerce')
        
        if pd.isna(cama_numeric) and pd.isna(expected_numeric):
            return True
        elif pd.isna(cama_numeric) or pd.isna(expected_numeric):
            return False
        else:
            return np.isclose(cama_numeric, expected_numeric, equal_nan=False, rtol=1e-9, atol=tolerance)
    except:
        return str(cama_val).strip().lower() == str(expected_cama).strip().lower()

def calculate_difference(val1, val2):
    """Calculate the difference between two values."""
    try:
        num1 = pd.to_numeric(val1, errors='raise')
        num2 = pd.to_numeric(val2, errors='raise')
        
        if pd.isna(num1) or pd.isna(num2):
            return "N/A"
        
        diff = num1 - num2
        return f"{diff:,.2f}"
    except (ValueError, TypeError):
        return "Text difference"

def compare_data_enhanced(df_mls, df_cama, unique_id_col, cols_to_compare_mapping,
                         cols_to_compare_sum=None, cols_to_compare_categorical=None,
                         tolerance=0.01, skip_zeros=True):
    """Compare MLS and CAMA dataframes."""
    
    mls_id_col_name = unique_id_col['mls_col']
    cama_id_col_name = unique_id_col['cama_col']
    
    df_mls_renamed = df_mls.copy()
    df_mls_renamed = df_mls_renamed.rename(columns={mls_id_col_name: cama_id_col_name})
    
    merged_df = pd.merge(df_mls_renamed, df_cama, on=cama_id_col_name, how='outer', indicator=True)
    
    missing_in_cama = []
    missing_in_mls = []
    value_mismatches = []
    perfect_matches = []
    
    for index, row in merged_df.iterrows():
        record_id = row.get(cama_id_col_name)
        merge_status = row.get('_merge')
        
        if merge_status == 'left_only':
            listing_num = row.get('Listing #', '')
            closed_date = row.get('Closed Date', '')
            missing_in_cama.append({
                'Parcel_ID': record_id,
                'Listing_Number': listing_num,
                'Closed_Date': closed_date
            })
            
        elif merge_status == 'right_only':
            missing_in_mls.append({'Parcel_ID': record_id})
            
        elif merge_status == 'both':
            listing_num = row.get('Listing #', '')
            salekey = row.get('SALEKEY', '')
            address = row.get(ADDRESS_COLUMNS.get('address', 'Address'), '')
            city = row.get(ADDRESS_COLUMNS.get('city', 'City'), '')
            state = row.get(ADDRESS_COLUMNS.get('state', 'State or Province'), '')
            zip_code = row.get(ADDRESS_COLUMNS.get('zip', 'Postal Code'), '')
            
            record_mismatches = []
            fields_compared = []
            
            # Standard 1-to-1 comparisons
            for mapping in cols_to_compare_mapping:
                mls_col = mapping['mls_col']
                cama_col = mapping['cama_col']
                
                if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                    continue
                
                mls_val = row.get(mls_col)
                cama_val = row.get(cama_col)
                
                mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')
                
                if mls_is_blank or cama_is_blank:
                    continue
                
                fields_compared.append(mls_col)
                
                if skip_zeros:
                    try:
                        mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                        cama_numeric = pd.to_numeric(cama_val, errors='coerce')
                        if (pd.notna(mls_numeric) and mls_numeric == 0) or (pd.notna(cama_numeric) and cama_numeric == 0):
                            continue
                    except:
                        pass
                
                if not values_equal(mls_val, cama_val, tolerance):
                    record_mismatches.append({
                        'Parcel_ID': record_id,
                        'Listing_Number': listing_num,
                        'SALEKEY': salekey,
                        'Address': address,
                        'City': city,
                        'State': state,
                        'Zip': zip_code,
                        'Field_MLS': mls_col,
                        'Field_CAMA': cama_col,
                        'MLS_Value': mls_val,
                        'CAMA_Value': cama_val,
                        'Difference': calculate_difference(mls_val, cama_val)
                    })
            
            # Sum comparisons
            if cols_to_compare_sum:
                for mapping in cols_to_compare_sum:
                    mls_col = mapping['mls_col']
                    cama_cols = mapping['cama_cols']
                    
                    if mls_col not in merged_df.columns:
                        continue
                    
                    missing_cols = [col for col in cama_cols if col not in merged_df.columns]
                    if missing_cols:
                        continue
                    
                    mls_val = row.get(mls_col)
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    
                    if mls_is_blank:
                        continue
                    
                    cama_sum = 0
                    all_cama_blank = True
                    for col in cama_cols:
                        val = row.get(col)
                        if pd.notna(val):
                            all_cama_blank = False
                            try:
                                cama_sum += pd.to_numeric(val, errors='coerce')
                            except:
                                pass
                    
                    if all_cama_blank:
                        continue
                    
                    fields_compared.append(mls_col)
                    
                    if skip_zeros:
                        try:
                            mls_numeric = pd.to_numeric(mls_val, errors='coerce')
                            if (pd.notna(mls_numeric) and mls_numeric == 0) or cama_sum == 0:
                                continue
                        except:
                            pass
                    
                    if not values_equal(mls_val, cama_sum, tolerance):
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': f"SUM({', '.join(cama_cols)})",
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_sum,
                            'Difference': calculate_difference(mls_val, cama_sum)
                        })
            
            # Categorical comparisons
            if cols_to_compare_categorical:
                for mapping in cols_to_compare_categorical:
                    mls_col = mapping['mls_col']
                    cama_col = mapping['cama_col']
                    
                    if mls_col not in merged_df.columns or cama_col not in merged_df.columns:
                        continue
                    
                    mls_val = row.get(mls_col)
                    cama_val = row.get(cama_col)
                    
                    mls_is_blank = pd.isna(mls_val) or (isinstance(mls_val, str) and mls_val.strip() == '')
                    cama_is_blank = pd.isna(cama_val) or (isinstance(cama_val, str) and cama_val.strip() == '')
                    
                    if mls_is_blank or cama_is_blank:
                        continue
                    
                    fields_compared.append(mls_col)
                    
                    if not categorical_match(mls_val, cama_val, mapping, tolerance):
                        check_text = mapping.get('mls_check_contains', '')
                        case_sensitive = mapping.get('case_sensitive', False)
                        mls_str = str(mls_val).strip() if pd.notna(mls_val) else ''
                        
                        if not case_sensitive:
                            mls_str = mls_str.lower()
                            check_text_lower = check_text.lower()
                        else:
                            check_text_lower = check_text
                        
                        text_found = check_text_lower in mls_str
                        expected_cama = mapping.get('cama_expected_if_true') if text_found else mapping.get('cama_expected_if_false')
                        
                        record_mismatches.append({
                            'Parcel_ID': record_id,
                            'Listing_Number': listing_num,
                            'SALEKEY': salekey,
                            'Address': address,
                            'City': city,
                            'State': state,
                            'Zip': zip_code,
                            'Field_MLS': mls_col,
                            'Field_CAMA': cama_col,
                            'MLS_Value': mls_val,
                            'CAMA_Value': cama_val,
                            'Expected_CAMA_Value': expected_cama,
                            'Match_Rule': f"If '{check_text}' in {mls_col}, then {cama_col} should be {mapping.get('cama_expected_if_true')}, else {mapping.get('cama_expected_if_false')}"
                        })
            
            if not record_mismatches and fields_compared:
                perfect_matches.append({
                    'Parcel_ID': record_id,
                    'Listing_Number': listing_num,
                    'SALEKEY': salekey,
                    'Address': address,
                    'City': city,
                    'State': state,
                    'Zip': zip_code,
                    'Fields_Compared': len(fields_compared),
                    'Fields_List': ', '.join(fields_compared)
                })
            
            value_mismatches.extend(record_mismatches)
    
    return (pd.DataFrame(missing_in_cama), pd.DataFrame(missing_in_mls),
            pd.DataFrame(value_mismatches), pd.DataFrame(perfect_matches))

def create_excel_with_hyperlinks(df, parcel_url_template):
    """Create Excel file with hyperlinks."""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Data']
    
    # Add Parcel_ID hyperlinks
    if 'Parcel_ID' in df.columns and parcel_url_template:
        parcel_col_idx = list(df.columns).index('Parcel_ID') + 1
        
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=parcel_col_idx)
            parcel_value = cell.value
            if parcel_value and str(parcel_value).strip():
                url = parcel_url_template.format(parcel_id=parcel_value)
                cell.hyperlink = url
                cell.style = 'Hyperlink'
    
    # Add Address hyperlinks to Zillow
    if 'Address' in df.columns:
        address_col_idx = list(df.columns).index('Address') + 1
        has_location_data = all(col in df.columns for col in ['City', 'Zip'])
        
        if has_location_data:
            city_col_idx = list(df.columns).index('City') + 1
            zip_col_idx = list(df.columns).index('Zip') + 1
            
            for row_idx in range(2, len(df) + 2):
                address_cell = ws.cell(row=row_idx, column=address_col_idx)
                address_value = address_cell.value
                
                if address_value and str(address_value).strip():
                    city = ws.cell(row=row_idx, column=city_col_idx).value
                    zip_code = ws.cell(row=row_idx, column=zip_col_idx).value
                    
                    url = format_zillow_url(address_value, city, 'OH', zip_code)
                    if url:
                        address_cell.hyperlink = url
                        address_cell.style = 'Hyperlink'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Main application logic
if mls_file and cama_file:
    try:
        # Load data
        with st.spinner('Loading data...'):
            df_mls = pd.read_excel(mls_file)
            df_cama = pd.read_excel(cama_file)
        
        st.success(f"✅ Loaded {len(df_mls)} MLS records and {len(df_cama)} CAMA records")
        
        # Show data previews
        with st.expander("📊 Preview MLS Data"):
            st.dataframe(df_mls.head())
        
        with st.expander("📊 Preview CAMA Data"):
            st.dataframe(df_cama.head())
        
        # Run comparison button
        if st.button("🔍 Run Comparison", type="primary"):
            with st.spinner('Comparing data...'):
                unique_id_col = {'mls_col': unique_id_mls, 'cama_col': unique_id_cama}
                
                df_missing_cama, df_missing_mls, df_value_mismatches, df_perfect_matches = compare_data_enhanced(
                    df_mls, df_cama, unique_id_col,
                    COLUMNS_TO_COMPARE,
                    cols_to_compare_sum=COLUMNS_TO_COMPARE_SUM,
                    cols_to_compare_categorical=COLUMNS_TO_COMPARE_CATEGORICAL,
                    tolerance=numeric_tolerance,
                    skip_zeros=skip_zero_values
                )
            
            # Display results
            st.header("📈 Results Summary")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("❌ Missing in CAMA", len(df_missing_cama))
            with col2:
                st.metric("❌ Missing in MLS", len(df_missing_mls))
            with col3:
                st.metric("⚠️ Value Mismatches", len(df_value_mismatches))
            with col4:
                st.metric("✅ Perfect Matches", len(df_perfect_matches))
            
            # Mismatches by field
            if not df_value_mismatches.empty:
                st.subheader("📊 Mismatches by Field")
                mismatch_counts = df_value_mismatches['Field_MLS'].value_counts()
                st.bar_chart(mismatch_counts)
            
            # Display and download options
            st.header("📥 Download Reports")
            
            if not df_missing_cama.empty:
                st.subheader("Missing in CAMA")
                st.dataframe(df_missing_cama)
                excel_data = create_excel_with_hyperlinks(df_missing_cama, parcel_url_template)
                st.download_button(
                    "⬇️ Download Missing in CAMA (Excel)",
                    excel_data,
                    "missing_in_CAMA.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if not df_missing_mls.empty:
                st.subheader("Missing in MLS")
                st.dataframe(df_missing_mls)
                excel_data = create_excel_with_hyperlinks(df_missing_mls, parcel_url_template)
                st.download_button(
                    "⬇️ Download Missing in MLS (Excel)",
                    excel_data,
                    "missing_in_MLS.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if not df_value_mismatches.empty:
                st.subheader("Value Mismatches")
                st.dataframe(df_value_mismatches)
                excel_data = create_excel_with_hyperlinks(df_value_mismatches, parcel_url_template)
                st.download_button(
                    "⬇️ Download Value Mismatches (Excel)",
                    excel_data,
                    "value_mismatches.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if not df_perfect_matches.empty:
                st.subheader("Perfect Matches")
                st.dataframe(df_perfect_matches)
                excel_data = create_excel_with_hyperlinks(df_perfect_matches, parcel_url_template)
                st.download_button(
                    "⬇️ Download Perfect Matches (Excel)",
                    excel_data,
                    "perfect_matches.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        st.exception(e)

else:
    st.info("👈 Please upload both MLS and CAMA Excel files to begin.")
    
    # Show instructions
    st.markdown("""
    ### 📖 Instructions
    
    1. **Upload Files**: Use the sidebar to upload your MLS and CAMA Excel files
    2. **Configure Settings**: Adjust column names, tolerance, and URL templates as needed
    3. **Run Comparison**: Click the "Run Comparison" button to analyze the data
    4. **Download Reports**: Download Excel reports with clickable hyperlinks
    
    ### 🔗 Hyperlinks
    
    - **Parcel ID**: Links to Stark County CAMA system
    - **Address**: Links to Zillow property search
    
    ### ✨ Features
    
    - Compare multiple fields between MLS and CAMA data
    - Handle sum comparisons (multiple CAMA fields)
    - Categorical text-to-numeric matching
    - Configurable tolerance for numeric comparisons
    - Option to skip zero values
    - Excel reports with clickable hyperlinks
    """)
